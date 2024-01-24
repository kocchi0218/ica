import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

# ヘッダー
st.title('生産スケジュールと点眼洗眼帳票分析アプリ')

def normalize_production_line_name(name):
    """
    Replace full-width brackets with half-width brackets in the production line name.
    """
    return name.replace('（', '(').replace('）', ')')


# ファイルアップロード
uploaded_file_icare = st.file_uploader("アイケア変バリスケジュール管理ファイルをアップロードしてください", type=['xlsx'])
uploaded_file_ueno = st.file_uploader("上野点眼洗眼帳票ファイルをアップロードしてください", type=['xlsx'])

if uploaded_file_icare is not None and uploaded_file_ueno is not None:
    # アイケア変バリスケジュール管理ファイルの読み込み
    icare_workbook = openpyxl.load_workbook(BytesIO(uploaded_file_icare.read()), data_only=True)
    icare_sheet = icare_workbook.active

    # Extract data for 生産略称、対象、スケール
    prod_abbrev_target_scale_items = []
    for row in icare_sheet.iter_rows(min_row=4, min_col=2, max_col=4, max_row=icare_sheet.max_row):
        prod_abbrev_value = row[0].value  # Column B (生産略称)
        target_value = row[1].value      # Column C (対象)
        scale_value = row[2].value       # Column D (スケール)
        if prod_abbrev_value is not None and target_value is not None and scale_value is not None:
            prod_abbrev_target_scale_items.append((prod_abbrev_value, target_value, scale_value))

    # Create DataFrame
    df_prod_abbrev_target_scale = pd.DataFrame(prod_abbrev_target_scale_items, columns=['生産略称', '対象', 'スケール'])

    # 上野点眼洗眼帳票ファイルの読み込み
    ueno_workbook = openpyxl.load_workbook(BytesIO(uploaded_file_ueno.read()), data_only=True)
    ueno_sheet = ueno_workbook.active

    # Set search conditions and execute search
    prod_abbrev_items_for_search_with_scale = df_prod_abbrev_target_scale[
        df_prod_abbrev_target_scale['対象'] == '調製'
    ]['生産略称'].tolist()

    prod_abbrev_items_for_search_exact_with_scale = [item + '調' for item in prod_abbrev_items_for_search_with_scale] + \
                                                    [item + '調製品' for item in prod_abbrev_items_for_search_with_scale] + \
                                                    [item + '(輸ろ' for item in prod_abbrev_items_for_search_with_scale]

    # Extract production line information
    production_line_info = {
        41: ueno_sheet.cell(row=4, column=41).value,  # Information in column 41
        44: ueno_sheet.cell(row=4, column=44).value,  # Information in column 44
        47: ueno_sheet.cell(row=4, column=47).value   # Information in column 47
    }

    # Extract search results with production line info
    exact_matches_with_A_col_with_scale_and_line = {}
    for row in ueno_sheet.iter_rows(min_row=1, max_row=ueno_sheet.max_row):
        for col_range_start, col_range_end, additional_col in [(41, 43, 42), (44, 46, 45), (47, 49, 48)]:
            for col in range(col_range_start, col_range_end + 1):
                cell_value = row[col-1].value
                if cell_value and cell_value in prod_abbrev_items_for_search_exact_with_scale:
                    line_info = production_line_info[col_range_start]
                    additional_data = row[additional_col-1].value  # Data from the additional column

                    match_info = (row[0].value, line_info, additional_data)
                    if cell_value in exact_matches_with_A_col_with_scale_and_line:
                        exact_matches_with_A_col_with_scale_and_line[cell_value].append(match_info)
                    else:
                        exact_matches_with_A_col_with_scale_and_line[cell_value] = [match_info]

    # Flatten the 'Match Info' into separate rows
    flattened_data = []
    for item, match_infos in exact_matches_with_A_col_with_scale_and_line.items():
        for date, line, additional in match_infos:
            flattened_data.append([item, date, line, additional])

    # Create a new DataFrame from the flattened data
    df_flattened = pd.DataFrame(flattened_data, columns=['Item', 'Date', 'Production Line', 'Additional Data'])

    # 生産ラインごとにデータフレームを分ける
    grouped_by_line = df_flattened.groupby('Production Line')

    # 各生産ラインのデータフレームを作成し、出力
    df_line_1 = grouped_by_line.get_group('点眼調製1号(B3・B4）') if '点眼調製1号(B3・B4）' in grouped_by_line.groups else pd.DataFrame()
    df_line_2 = grouped_by_line.get_group('点眼調製2号(B7・B8）') if '点眼調製2号(B7・B8）' in grouped_by_line.groups else pd.DataFrame()
    df_line_3 = grouped_by_line.get_group('点眼調製3号(B0）') if '点眼調製3号(B0）' in grouped_by_line.groups else pd.DataFrame()

    # 生産ラインごとのデータフレームを表示
    st.header("点眼調製1号(B3・B4）のデータ:")
    st.dataframe(df_line_1)

    st.header("点眼調製2号(B7・B8）のデータ:")
    st.dataframe(df_line_2)

    st.header("点眼調製3号(B0）のデータ:")
    st.dataframe(df_line_3)